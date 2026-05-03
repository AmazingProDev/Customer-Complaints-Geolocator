#!/usr/bin/env python3

import argparse
import math
import re
import subprocess
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


MOROCCO_LON_MIN = -17.5
MOROCCO_LON_MAX = -0.5
MOROCCO_LAT_MIN = 20.0
MOROCCO_LAT_MAX = 36.5
GDALTRANSFORM = "/opt/homebrew/bin/gdaltransform"
DATA_START_ROW = 8
SOURCE_X_COL = 6
SOURCE_Y_COL = 7
APPENDED_HEADERS = [
    "source_coord_type",
    "normalized_longitude",
    "normalized_latitude",
    "normalization_action",
    "normalization_confidence",
    "needs_review",
    "review_reason",
]


@dataclass
class RowNormalization:
    row_number: int
    code: str | None
    source_coord_type: str
    normalized_longitude: float | None
    normalized_latitude: float | None
    normalization_action: str
    normalization_confidence: float
    needs_review: bool
    review_reason: str
    projected_x: float | None = None
    projected_y: float | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize ANRT coordinate columns to WGS84 decimal longitude/latitude."
    )
    parser.add_argument("--input", required=True, help="Input workbook path")
    parser.add_argument("--sheet", default="IAM", help="Worksheet name to process")
    parser.add_argument("--output", required=True, help="Output workbook path")
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def is_blank(value: object) -> bool:
    return normalize_text(value) == ""


def parse_numeric_value(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    compact = text.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", compact):
        return float(compact)
    return None


def in_morocco_bbox(lon: float, lat: float) -> bool:
    return (
        MOROCCO_LON_MIN <= lon <= MOROCCO_LON_MAX
        and MOROCCO_LAT_MIN <= lat <= MOROCCO_LAT_MAX
    )


def looks_like_dms(value: object) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    return any(marker in text for marker in ("°", "'", '"'))


def cleanup_dms_component(component: str) -> str:
    component = component.strip().replace(",", ".")
    component = component.replace("S", "5").replace("s", "5")
    component = component.replace("O", "0").replace("o", "0")
    return component


def parse_dms_value(value: object, axis: str) -> tuple[float | None, str]:
    text = normalize_text(value)
    if not text:
        return None, "empty_dms"

    raw = text.replace(" ", "")
    direction = None

    while raw and raw[-1] in "\"'":
        raw = raw[:-1]

    if raw:
        tail = raw[-1].upper()
        if tail in {"N", "S", "E", "W", "O", "0"}:
            direction = "O" if tail == "0" else tail
            raw = raw[:-1]
            while raw and raw[-1] in "\"'":
                raw = raw[:-1]

    normalized = raw.replace("°", "'").replace('"', "'")
    parts = [part for part in normalized.split("'") if part]

    if len(parts) == 4 and len(parts[-1]) == 1 and parts[-1].upper() in {"N", "S", "E", "W", "O", "0"}:
        direction = "O" if parts[-1].upper() == "0" else parts[-1].upper()
        parts = parts[:-1]

    if len(parts) != 3:
        return None, "invalid_dms_token_count"

    deg_text, min_text, sec_text = (cleanup_dms_component(part) for part in parts)

    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?", deg_text):
        return None, "invalid_dms_degrees"
    if not re.fullmatch(r"\d+(?:\.\d+)?", min_text):
        return None, "invalid_dms_minutes"
    if not re.fullmatch(r"\d+(?:\.\d+)?", sec_text):
        return None, "invalid_dms_seconds"

    degrees = float(deg_text)
    minutes = float(min_text)
    seconds = float(sec_text)

    if minutes >= 60 or seconds >= 60:
        return None, "invalid_dms_range"

    absolute = abs(degrees) + minutes / 60.0 + seconds / 3600.0

    if direction is None:
        direction = "O" if axis == "lon" else "N"

    direction = direction.upper()
    sign = 1.0
    if direction in {"W", "O", "S"}:
        sign = -1.0

    value_decimal = round(sign * absolute, 9)
    return value_decimal, "parsed_dms"


def classify_coordinate_type(raw_x: object, raw_y: object, parsed_x: float | None, parsed_y: float | None) -> str:
    if is_blank(raw_x) and is_blank(raw_y):
        return "blank"
    if looks_like_dms(raw_x) or looks_like_dms(raw_y):
        return "sexagesimal_or_text"
    if parsed_x is None or parsed_y is None:
        return "sexagesimal_or_text"
    if abs(parsed_x) > 1000 or abs(parsed_y) > 1000:
        return "projected_zone_i"
    return "decimal"


def normalize_decimal_pair(
    x: float,
    y: float,
    raw_x: object,
    raw_y: object,
) -> RowNormalization:
    if in_morocco_bbox(x, y):
        return RowNormalization(
            row_number=0,
            code=None,
            source_coord_type="decimal",
            normalized_longitude=round(x, 7),
            normalized_latitude=round(y, 7),
            normalization_action="accepted_decimal_as_is",
            normalization_confidence=1.0,
            needs_review=False,
            review_reason="",
        )

    if 0 <= x <= 20 and MOROCCO_LAT_MIN <= y <= MOROCCO_LAT_MAX:
        lon = -abs(x)
        lat = y
        return RowNormalization(
            row_number=0,
            code=None,
            source_coord_type="decimal",
            normalized_longitude=round(lon, 7),
            normalized_latitude=round(lat, 7),
            normalization_action="fixed_missing_longitude_minus",
            normalization_confidence=0.98,
            needs_review=False,
            review_reason="",
        )

    if MOROCCO_LAT_MIN <= x <= MOROCCO_LAT_MAX and MOROCCO_LON_MIN <= y <= MOROCCO_LON_MAX:
        return RowNormalization(
            row_number=0,
            code=None,
            source_coord_type="decimal",
            normalized_longitude=round(y, 7),
            normalized_latitude=round(x, 7),
            normalization_action="swapped_lat_lon_columns",
            normalization_confidence=0.97,
            needs_review=False,
            review_reason="",
        )

    if MOROCCO_LAT_MIN <= x <= MOROCCO_LAT_MAX and 0 <= y <= 20:
        lon = -abs(y)
        lat = x
        return RowNormalization(
            row_number=0,
            code=None,
            source_coord_type="decimal",
            normalized_longitude=round(lon, 7),
            normalized_latitude=round(lat, 7),
            normalization_action="swapped_columns_and_fixed_missing_minus",
            normalization_confidence=0.95,
            needs_review=False,
            review_reason="",
        )

    repaired = try_decimal_point_repair(x, y, raw_x, raw_y)
    if repaired is not None:
        return repaired

    return RowNormalization(
        row_number=0,
        code=None,
        source_coord_type="decimal",
        normalized_longitude=None,
        normalized_latitude=None,
        normalization_action="flagged_decimal_outlier",
        normalization_confidence=0.0,
        needs_review=True,
        review_reason="decimal_values_do_not_match_conservative_morocco_rules",
    )


def decimal_candidates(value: float, raw_value: object) -> list[float]:
    candidates: list[float] = []
    text = normalize_text(raw_value)
    text_has_decimal = "." in text or "," in text

    if text and not text_has_decimal and abs(value) >= 100:
        for shift in range(1, 4):
            shifted = value / (10 ** shift)
            if abs(shifted) < 100:
                candidates.append(shifted)

    deduped: list[float] = []
    seen = set()
    for candidate in candidates:
        rounded = round(candidate, 9)
        if rounded not in seen:
            seen.add(rounded)
            deduped.append(candidate)
    return deduped


def try_decimal_point_repair(
    x: float,
    y: float,
    raw_x: object,
    raw_y: object,
) -> RowNormalization | None:
    candidates: list[tuple[float, float, str]] = []

    for repaired_x in decimal_candidates(x, raw_x):
        if 0 <= repaired_x <= 20 and MOROCCO_LAT_MIN <= y <= MOROCCO_LAT_MAX:
            candidates.append(
                (-abs(repaired_x), y, "repaired_decimal_point_in_x_and_fixed_missing_minus")
            )

    for repaired_y in decimal_candidates(y, raw_y):
        if 0 <= x <= 20 and MOROCCO_LAT_MIN <= repaired_y <= MOROCCO_LAT_MAX:
            candidates.append(
                (-abs(x), repaired_y, "repaired_decimal_point_in_y_and_fixed_missing_minus")
            )

    unique_candidates: dict[tuple[float, float], str] = {}
    for lon, lat, action in candidates:
        if in_morocco_bbox(lon, lat):
            unique_candidates[(round(lon, 7), round(lat, 7))] = action

    if len(unique_candidates) != 1:
        return None

    (lon, lat), action = next(iter(unique_candidates.items()))
    return RowNormalization(
        row_number=0,
        code=None,
        source_coord_type="decimal",
        normalized_longitude=lon,
        normalized_latitude=lat,
        normalization_action=action,
        normalization_confidence=0.75,
        needs_review=False,
        review_reason="",
    )


def run_projected_conversion(pairs: Iterable[tuple[float, float]]) -> list[tuple[float | None, float | None, str]]:
    pair_list = list(pairs)
    if not pair_list:
        return []

    input_lines = "\n".join(f"{x} {y}" for x, y in pair_list) + "\n"
    completed = subprocess.run(
        [
            GDALTRANSFORM,
            "-s_srs",
            "EPSG:26191",
            "-t_srs",
            "EPSG:4326",
        ],
        input=input_lines,
        text=True,
        capture_output=True,
        check=False,
    )

    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or "Projected conversion failed")

    lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    if len(lines) != len(pair_list):
        raise RuntimeError(
            f"Projected conversion returned {len(lines)} rows for {len(pair_list)} inputs"
        )

    converted: list[tuple[float | None, float | None, str]] = []
    for line in lines:
        parts = line.split()
        if len(parts) < 2:
            converted.append((None, None, "invalid_projected_output"))
            continue
        lon = float(parts[0])
        lat = float(parts[1])
        converted.append((round(lon, 7), round(lat, 7), "converted_projected_zone_i"))
    return converted


def normalize_row(row_number: int, code: str | None, raw_x: object, raw_y: object) -> RowNormalization:
    parsed_x = parse_numeric_value(raw_x)
    parsed_y = parse_numeric_value(raw_y)
    source_coord_type = classify_coordinate_type(raw_x, raw_y, parsed_x, parsed_y)

    if source_coord_type == "blank":
        return RowNormalization(
            row_number=row_number,
            code=code,
            source_coord_type="blank",
            normalized_longitude=None,
            normalized_latitude=None,
            normalization_action="left_blank",
            normalization_confidence=1.0,
            needs_review=False,
            review_reason="",
        )

    if source_coord_type == "sexagesimal_or_text":
        dms_x, reason_x = parse_dms_value(raw_x, "lon")
        dms_y, reason_y = parse_dms_value(raw_y, "lat")
        if dms_x is not None and dms_y is not None:
            lon = round(dms_x, 7)
            lat = round(dms_y, 7)
            needs_review = not in_morocco_bbox(lon, lat)
            return RowNormalization(
                row_number=row_number,
                code=code,
                source_coord_type="sexagesimal_or_text",
                normalized_longitude=lon,
                normalized_latitude=lat,
                normalization_action="parsed_dms_with_ocr_cleanup",
                normalization_confidence=0.85 if not needs_review else 0.45,
                needs_review=needs_review,
                review_reason="" if not needs_review else "parsed_dms_result_outside_morocco_bbox",
            )

        if parsed_x is not None and parsed_y is not None and (abs(parsed_x) > 1000 or abs(parsed_y) > 1000):
            return RowNormalization(
                row_number=row_number,
                code=code,
                source_coord_type="projected_zone_i",
                normalized_longitude=None,
                normalized_latitude=None,
                normalization_action="queued_projected_conversion",
                normalization_confidence=0.9,
                needs_review=False,
                review_reason="",
                projected_x=parsed_x,
                projected_y=parsed_y,
            )

        if parsed_x is not None and parsed_y is not None:
            decimal_result = normalize_decimal_pair(parsed_x, parsed_y, raw_x, raw_y)
            decimal_result.row_number = row_number
            decimal_result.code = code
            decimal_result.source_coord_type = "decimal"
            return decimal_result

        return RowNormalization(
            row_number=row_number,
            code=code,
            source_coord_type="sexagesimal_or_text",
            normalized_longitude=None,
            normalized_latitude=None,
            normalization_action="flagged_unparsed_text_coordinates",
            normalization_confidence=0.0,
            needs_review=True,
            review_reason=f"could_not_parse_coordinates:{reason_x}|{reason_y}",
        )

    if source_coord_type == "projected_zone_i":
        return RowNormalization(
            row_number=row_number,
            code=code,
            source_coord_type="projected_zone_i",
            normalized_longitude=None,
            normalized_latitude=None,
            normalization_action="queued_projected_conversion",
            normalization_confidence=0.9,
            needs_review=False,
            review_reason="",
            projected_x=parsed_x,
            projected_y=parsed_y,
        )

    decimal_result = normalize_decimal_pair(parsed_x, parsed_y, raw_x, raw_y)
    decimal_result.row_number = row_number
    decimal_result.code = code
    return decimal_result


def write_row_metadata(worksheet, row_number: int, first_metadata_col: int, result: RowNormalization) -> None:
    values = [
        result.source_coord_type,
        result.normalized_longitude,
        result.normalized_latitude,
        result.normalization_action,
        result.normalization_confidence,
        "yes" if result.needs_review else "no",
        result.review_reason,
    ]
    for offset, value in enumerate(values):
        worksheet.cell(row=row_number, column=first_metadata_col + offset, value=value)


def rebuild_sheet_headers(worksheet, first_metadata_col: int) -> None:
    for offset, header in enumerate(APPENDED_HEADERS):
        worksheet.cell(row=6, column=first_metadata_col + offset, value=header)
        worksheet.cell(row=7, column=first_metadata_col + offset, value=header)


def replace_sheet(workbook, title: str):
    if title in workbook.sheetnames:
        existing = workbook[title]
        workbook.remove(existing)
    return workbook.create_sheet(title)


def write_audit_sheet(workbook, rows: list[RowNormalization]) -> None:
    sheet = replace_sheet(workbook, "Coordinate Audit")
    sheet.append(["Metric", "Value"])
    sheet.append(["total_rows", len(rows)])
    sheet.append(["review_rows", sum(1 for row in rows if row.needs_review)])
    sheet.append([])

    sheet.append(["source_coord_type", "count"])
    for key, count in Counter(row.source_coord_type for row in rows).most_common():
        sheet.append([key, count])
    sheet.append([])

    sheet.append(["normalization_action", "count"])
    for key, count in Counter(row.normalization_action for row in rows).most_common():
        sheet.append([key, count])
    sheet.append([])

    sheet.append(["review_reason", "count"])
    review_counter = Counter(row.review_reason for row in rows if row.review_reason)
    for key, count in review_counter.most_common():
        sheet.append([key, count])


def write_review_sheet(workbook, source_sheet_name: str, rows: list[RowNormalization]) -> None:
    source_sheet = workbook[source_sheet_name]
    sheet = replace_sheet(workbook, "Coordinate Review")

    header = ["original_row_number"]
    for column in range(1, 13):
        header.append(source_sheet.cell(row=6, column=column).value or f"col_{column}")
    header.extend(APPENDED_HEADERS)
    sheet.append(header)

    for row in rows:
        if not row.needs_review:
            continue
        record = [row.row_number]
        for column in range(1, 13):
            record.append(source_sheet.cell(row=row.row_number, column=column).value)
        record.extend(
            [
                row.source_coord_type,
                row.normalized_longitude,
                row.normalized_latitude,
                row.normalization_action,
                row.normalization_confidence,
                "yes",
                row.review_reason,
            ]
        )
        sheet.append(record)


def process_workbook(input_path: Path, sheet_name: str, output_path: Path) -> tuple[list[RowNormalization], Path]:
    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]
    first_metadata_col = 13
    rebuild_sheet_headers(sheet, first_metadata_col)

    row_results: list[RowNormalization] = []
    projected_queue: list[tuple[int, float, float]] = []

    for row_number in range(DATA_START_ROW, sheet.max_row + 1):
        code = sheet.cell(row=row_number, column=1).value
        raw_x = sheet.cell(row=row_number, column=SOURCE_X_COL).value
        raw_y = sheet.cell(row=row_number, column=SOURCE_Y_COL).value
        result = normalize_row(row_number, code, raw_x, raw_y)
        if result.projected_x is not None and result.projected_y is not None:
            projected_queue.append((len(row_results), result.projected_x, result.projected_y))
        row_results.append(result)

    converted = run_projected_conversion((x, y) for _, x, y in projected_queue)
    for (result_index, projected_x, projected_y), (lon, lat, action) in zip(projected_queue, converted):
        result = row_results[result_index]
        if lon is None or lat is None:
            result.normalization_action = "projected_conversion_failed"
            result.normalization_confidence = 0.0
            result.needs_review = True
            result.review_reason = "projected_conversion_failed"
        else:
            result.normalized_longitude = lon
            result.normalized_latitude = lat
            result.normalization_action = action
            if in_morocco_bbox(lon, lat):
                result.normalization_confidence = 0.93
                result.needs_review = False
                result.review_reason = ""
            else:
                result.normalization_confidence = 0.4
                result.needs_review = True
                result.review_reason = "projected_result_outside_morocco_bbox"

    for result in row_results:
        write_row_metadata(sheet, result.row_number, first_metadata_col, result)

    write_audit_sheet(workbook, row_results)
    write_review_sheet(workbook, sheet_name, row_results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return row_results, output_path


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    rows, saved_path = process_workbook(input_path, args.sheet, output_path)
    review_count = sum(1 for row in rows if row.needs_review)
    print(f"Saved normalized workbook to {saved_path}")
    print(f"Processed {len(rows)} rows; {review_count} rows flagged for review.")


if __name__ == "__main__":
    main()
